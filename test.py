import sys
import json
from PyQt6.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox, QMainWindow, QTextEdit, QDialog  # type: ignore
from openpyxl import load_workbook, Workbook  # type: ignore
import os
import qrcode  # type: ignore
from PIL import Image
from PIL.ImageQt import ImageQt
from PyQt6.QtGui import QPixmap  # type: ignore
from icecream import ic  # type: ignore


class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle('Login Window')
        self.setGeometry(100, 100, 400, 200)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        login_label = QLabel('Логин:  ')
        password_label = QLabel('Пароль:')

        self.login_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)

        login_layout = QHBoxLayout()
        login_layout.addWidget(login_label)
        login_layout.addWidget(self.login_edit)

        password_layout = QHBoxLayout()
        password_layout.addWidget(password_label)
        password_layout.addWidget(self.password_edit)

        layout.addLayout(login_layout)
        layout.addLayout(password_layout)

        auth_button = QPushButton('Авторизация')
        auth_button.clicked.connect(self.authenticate)

        register_button = QPushButton('Регистрация')
        register_button.clicked.connect(self.register_user)

        layout.addWidget(auth_button)
        layout.addWidget(register_button)

        self.setLayout(layout)

    def authenticate(self):
        ic.configureOutput(prefix='Auth| ')
        username = self.login_edit.text()
        password = self.password_edit.text()

        try:
            with open('Auth/users.json', 'r') as file:
                users = json.load(file)

            if username in users and users[username] == password:
                ic(username)
                self.show_second_window()
            else:
                self.show_error_message('Ошибка авторизации', 'Неверный логин или пароль')

        except FileNotFoundError:
            self.show_error_message('Ошибка', 'Файл с пользователями не найден')
        except json.JSONDecodeError:
            self.show_error_message('Ошибка', 'Ошибка при чтении файла с пользователями')

    def register_user(self):
        ic.configureOutput(prefix='Registr| ')
        username = self.login_edit.text()
        password = self.password_edit.text()

        try:
            with open('Auth/users.json', 'r') as file:
                users = json.load(file)


        except FileNotFoundError:
            users = {}

        users[username] = password
        if not os.path.exists('Auth/'):
            os.makedirs('Auth/')

        try:
            with open('Auth/users.json', 'w') as file:
                json.dump(users, file, indent=4)

            self.show_message('Успех', 'Пользователь успешно зарегистрирован')
            ic(users)

        except Exception as e:
            self.show_error_message('Ошибка', f'Ошибка при записи в файл: {str(e)}')
            ic(f'Ошибка при записи в файл: {str(e)}')

    def show_error_message(self, title, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec()

    def show_message(self, title, message):
        msg_box = QMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec()

    def show_second_window(self):
        second_window.show()
        self.hide()


class SecondWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle('Second Window')
        self.setGeometry(100, 100, 600, 400)

        self.init_ui()

    def init_ui(self):
        ic.configureOutput(prefix='Init2| ')
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()

        # Проверка наличия папки и файла
        database_folder = 'Database'
        database_file = 'data.xlsx'

        if not os.path.exists(database_folder):
            os.makedirs(database_folder)
            ic(f'Made {database_folder}')
        else:
            ic(f'Already exists {database_folder}')

        database_path = os.path.join(database_folder, database_file)
        if not os.path.exists(database_path):
            workbook = Workbook()
            workbook.save(database_path)
            ic(f'Made {database_path}')
        else:
            ic(f'Already exists {database_path}')

        view_db_button = QPushButton('Посмотреть содержимое БД')
        view_db_button.clicked.connect(self.view_database)

        add_record_button = QPushButton('Внести записи в БД')
        add_record_button.clicked.connect(self.add_record_to_database)

        view_qr_button = QPushButton('Открыть QR-код')
        view_qr_button.clicked.connect(self.view_qr_code)

        self.result_display = QTextEdit()

        layout.addWidget(view_db_button)
        layout.addWidget(add_record_button)
        layout.addWidget(view_qr_button)
        layout.addWidget(self.result_display)

        central_widget.setLayout(layout)

    def view_database(self):
        try:
            workbook = load_workbook('Database/data.xlsx')
            sheet = workbook.active

            data = []
            header = ['1 поле', '2 поле', '3 поле']

            # Добавление подписей колонок
            data.append('<p style="margin:0;">{}</p>'.format('\t'.join(map(str, header))))

            # Чередование цветов строк
            row_color = True

            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row_color:
                    data.append(
                        "<p style='background-color: #f2f2f2; margin:0;'>{}</p>".format('\t'.join(map(str, row))))
                else:
                    data.append("<p style='margin:0;'>{}</p>".format('\t'.join(map(str, row))))
                row_color = not row_color

            self.result_display.setHtml(''.join(data))

        except FileNotFoundError:
            self.show_error_message('Ошибка', 'Файл с базой данных не найден')
        except Exception as e:
            self.show_error_message('Ошибка', f'Ошибка при чтении базы данных: {str(e)}')

    def view_qr_code(self):
        ic.configureOutput(prefix='ViewQR| ')
        selected_row = self.result_display.textCursor().blockNumber()
        if selected_row < 0:
            return

        try:
            ic(selected_row)
            workbook = load_workbook('Database/data.xlsx')
            sheet = workbook.active

            qr_img_path = sheet.cell(row=ic(selected_row + 1), column=ic(sheet.max_column)).value
            ic(qr_img_path)
            if qr_img_path:
                qr_dialog = QRCodeDialog(qr_img_path, self)
                qr_dialog.exec()

        except FileNotFoundError:
            self.show_error_message('Ошибка', 'Файл с базой данных не найден')
        except Exception as e:
            self.show_error_message('Ошибка', f'Ошибка при чтении базы данных: {str(e)}')

    def add_record_to_database(self):
        add_record_dialog = AddRecordDialog(self)
        add_record_dialog.exec()

    def show_error_message(self, title, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec()


class AddRecordDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)

        self.setWindowTitle('Добавить запись в БД')
        self.setGeometry(200, 200, 400, 200)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        label1 = QLabel('Поле 1:')
        label2 = QLabel('Поле 2:')
        label3 = QLabel('Поле 3:')

        self.field1_edit = QLineEdit()
        self.field2_edit = QLineEdit()
        self.field3_edit = QLineEdit()

        layout.addWidget(label1)
        layout.addWidget(self.field1_edit)
        layout.addWidget(label2)
        layout.addWidget(self.field2_edit)
        layout.addWidget(label3)
        layout.addWidget(self.field3_edit)

        add_button = QPushButton('Добавить запись')
        add_button.clicked.connect(self.add_record)

        layout.addWidget(add_button)

        self.setLayout(layout)

    def add_record(self):
        field1 = self.field1_edit.text()
        field2 = self.field2_edit.text()
        field3 = self.field3_edit.text()

        try:
            workbook = load_workbook('Database/data.xlsx')
            sheet = workbook.active

            # Генерация QR-кода
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(f'{field1} | {field2} | {field3}')
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")

            # Сохранение QR-кода в файл
            qr_img_path = f'Database/qrcodes/{field1}_{field2}_{field3}_qrcode.png'
            qr_img.save(qr_img_path)

            # Добавление записи в базу данных
            sheet.append([field1, field2, field3, qr_img_path])
            workbook.save('Database/data.xlsx')

            self.parent().view_database()
            self.accept()

        except FileNotFoundError:
            self.parent().show_error_message('Ошибка', 'Файл с базой данных не найден')
        except Exception as e:
            self.parent().show_error_message('Ошибка', f'Ошибка при записи в базу данных: {str(e)}')


class QRCodeDialog(QDialog):
    def __init__(self, qr_img_path, parent):
        super().__init__(parent)

        self.setWindowTitle('QR-код')
        self.setGeometry(200, 200, 300, 300)

        self.init_ui(qr_img_path)

    def init_ui(self, qr_img_path):
        layout = QVBoxLayout()

        img_label = QLabel()
        img = Image.open(qr_img_path)
        img_label.setPixmap(QPixmap.fromImage(ImageQt(img)))

        layout.addWidget(img_label)

        self.setLayout(layout)


def main():
    app = QApplication(sys.argv)

    global second_window
    second_window = SecondWindow()

    login_window = LoginWindow()
    login_window.show()

    sys.exit(app.exec())


if __name__ == '__main__':
    main()
